from datetime import datetime

import openpyxl
import csv

from openpyxl.styles import numbers
from openpyxl.styles.numbers import NumberFormat

# 创建excel表格以及工作簿
wb = openpyxl.Workbook()
wb[wb.sheetnames[0]].title = '北邮'
wb.create_sheet('西电')

# 写北邮部分
BEIYOU_file = '../BEIYOU_1.csv'
with open(BEIYOU_file, 'r', encoding='utf-8') as bf:
    BEIYOU_reader = csv.DictReader(bf)
    i = 1
    sheet = wb['北邮']
    # 写表头
    head = ['序号', '招聘主题', '发布日期', '浏览次数']
    sheet.append(head)
    for row in BEIYOU_reader:
        # data用于保存要写入表格的数据
        data = list()
        # 分别在data中加入根据要求预处理后的序号、招聘主题、发布日期、浏览次数
        data.append(i)
        data.append(row['job_title'][2:-2].strip())
        data.append(row['job_date'][2:-2])
        data.append(0 if not row['job_views'] else int(row['job_views'][2:-2]))
        i = i + 1
        # 向表格写入数据
        for col in range(1, 5):
            sheet.cell(i, col).value = data[col - 1]
            if col == 3:
                # 设置格式为日期
                sheet.cell(i, col).number_format = numbers.FORMAT_DATE_YYYYMMDD2 + ';@'
            elif col == 4:
                # 设置格式为整数
                sheet.cell(i, col).number_format = numbers.FORMAT_NUMBER
            else:
                # 设置格式为文本
                sheet.cell(i, col).number_format = numbers.FORMAT_TEXT

# 写西电部分
XIDIAN_file = '../XIDIAN_1.csv'
with open(XIDIAN_file, 'r', encoding='utf-8') as xf:
    XIDIAN_reader = csv.DictReader(xf)
    i = 1
    sheet = wb['西电']
    head = ['序号', '招聘主题', '发布日期', '浏览次数']
    sheet.append(head)
    for row in XIDIAN_reader:
        data = list()
        data.append(i)
        data.append(row['job_title'][2:-2].strip())
        # 西电的时间包含时分，根据要求只截取到日期
        data.append(row['job_date'][2:12])
        data.append(0 if not row['job_views'] else int(row['job_views'][2:-2]))
        i = i + 1
        for col in range(1, 5):
            sheet.cell(i, col).value = data[col - 1]
            if col == 3:
                sheet.cell(i, col).number_format = numbers.FORMAT_DATE_YYYYMMDD2 + ';@'
            elif col == 4:
                sheet.cell(i, col).number_format = numbers.FORMAT_NUMBER
            else:
                sheet.cell(i, col).number_format = numbers.FORMAT_TEXT

# 保存表格并关闭
wb.save('./date.xlsx')
wb.close()
